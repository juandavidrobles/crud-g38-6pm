from openpyxl import load_workbook

filename = './data.xlsx'
file = load_workbook(filename)

# Read
def read_todos()->list[dict]:
  todos = []
  sheet = file['BD']
  rows = list(sheet.rows)
  for row in rows[1:]:
    todo = convert_row_to_dict(row)
    todos.append(todo)

  return todos

def convert_row_to_dict(row: tuple)->dict:
  dictionary = {
    'Id': row[0].value,
    'Tarea': row[1].value,
    'Descripcion': row[2].value,
    'Estado': row[3].value,
    'Fecha inicio': row[4].value,
    'Fecha finalizacion': row[5].value,
  }
  return dictionary

# todos = read_todos()
# for todo in todos:
#   print('---------------------------')
#   print(f"Tarea: {todo['Tarea']}\nDescripcion: {todo['Descripcion']}\nEstado: {todo['Estado']}")

# Create
def create_todo(todo: dict):
  sheet = file['BD']
  row_idx = sheet.max_row + 1
  sheet.cell(column=1, row=row_idx).value = todo.get('Id')
  sheet.cell(column=2, row=row_idx).value = todo.get('Tarea')
  sheet.cell(column=3, row=row_idx).value = todo.get('Descripcion')
  sheet.cell(column=4, row=row_idx).value = todo.get('Estado')
  sheet.cell(column=5, row=row_idx).value = todo.get('Fecha inicio')
  sheet.cell(column=6, row=row_idx).value = todo.get('Fecha finalizacion')

  file.save(filename)

# create_todo({
#   'Id': 5,
#   'Tarea': 'Desayunar',
#   'Descripcion': 'Comer el desayuno'
# })

# Delete
def delete_todo(id: int):
  sheet = file['BD']
  for row in sheet.rows:
    if (row[0].value == id):
      idx = row[0].row
      sheet.delete_rows(idx)

  file.save(filename)

# delete_todo(4)

def find_idx_by_todo_id(sheet, id: int)-> int:
  for row in sheet.rows:
    if (row[0].value == id):
      return row[0].row

# Update
def update_todo(id: int, todo: dict):
  sheet = file['BD']
  idx = find_idx_by_todo_id(sheet, id)
  if (todo.get('Tarea')): sheet.cell(row=idx, column=2).value = todo.get('Tarea')
  if (todo.get('Descripcion')): sheet.cell(row=idx, column=3).value = todo.get('Descripcion')
  if (todo.get('Estado')): sheet.cell(row=idx, column=4).value = todo.get('Estado')
  if (todo.get('Fecha inicio')): sheet.cell(row=idx, column=5).value = todo.get('Fecha inicio')
  if (todo.get('Fecha finalizacion')): sheet.cell(row=idx, column=6).value = todo.get('Fecha finalizacion')

  file.save(filename)

update_todo(3, {'Tarea': 'Dormir'})
